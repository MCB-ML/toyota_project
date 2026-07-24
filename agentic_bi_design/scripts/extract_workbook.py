"""
P0 — Workbook Inventory extractor.

Reads server/schema/ktws_측정값_쿼리화.xlsx (READ-ONLY, never modified) and
produces normalized JSON catalogs under agentic_bi_design/inventory/.

Header layout differs per sheet (columns shift, some sheets add/drop a
column) — see agentic_bi_design/docs/known_issues.md item KI-1. We resolve
columns by label match, not fixed position, and fall back to raw capture
for anything we can't confidently classify (nothing is silently dropped).
"""
import json
import hashlib
import re
import unicodedata
from pathlib import Path

import openpyxl

SRC_XLSX = Path(r"c:/Project/toyota_project/toyota_project/louis/dashboard/server/schema/ktws_측정값_쿼리화.xlsx")
OUT_DIR = Path(r"c:/Project/toyota_project/toyota_project/agentic_bi_design/inventory")

# label -> canonical field. Matched by "normalized header contains this substring".
# order matters: more specific keys first so "시각적 개체명" doesn't get
# swallowed by the generic "시각적 개체" substring match.
HEADER_SYNONYMS = {
    "page": ["페이지"],
    "value_label": ["개체값이름", "개체 값 이름"],
    "visual_name": ["시각적개체명", "시각적 개체명", "시각적개체 명", "시각적 개체 명"],
    "visual_type": ["시각적개체", "시각적 개체"],
    "used_measures": ["사용측정값", "사용 측정값"],
    "related_measures": ["관련측정값", "관련 측정값"],
    "dax_expression": ["DAX식", "DAX 식"],
    "query_candidate": ["GOLD쿼리", "GOLD 쿼리", "GOLDSQL", "GOLD SQL", "쿼리"],
    "business_meaning": ["업무의미", "업무 의미"],
    "slicers_filters": ["슬라이서", "필터"],
}
FIELD_ORDER = ["page", "value_label", "visual_name", "visual_type", "used_measures",
               "related_measures", "dax_expression", "query_candidate",
               "business_meaning", "slicers_filters"]


def query_priority(header_label):
    """Sheets carry multiple draft/merged/final GOLD-query columns (draft ->
    합본 -> 최종). Rank by header wording so the FINAL one always wins,
    regardless of its column position (position isn't consistent — one
    sheet even puts '최종 GOLD 쿼리' to the left of '합본')."""
    n = norm(header_label)
    if "최종통합" in n or "최종 통합" in norm(header_label):
        return 100
    if "최종" in n and "v2" in n.lower():
        return 95
    if "최종" in n and "v1" in n.lower():
        return 90
    if "최종" in n:
        return 85
    if "합본" in n:
        return 70
    if "gold" in n.lower():
        return 50
    return 10


def norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFC", str(s))
    return re.sub(r"\s+", "", s).strip()


def classify_header(cell_text):
    n = norm(cell_text)
    if not n:
        return None
    for field in FIELD_ORDER:
        for syn in HEADER_SYNONYMS[field]:
            if norm(syn) in n:
                return field
    return None


def find_header_row(ws, max_scan=5):
    """Header isn't always row 1 in every sheet — scan first N rows, pick
    the one with the most classifiable columns."""
    best_row, best_score = 1, -1
    for r in range(1, min(max_scan, ws.max_row) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        score = sum(1 for v in row_vals if classify_header(v))
        if score > best_score:
            best_score, best_row = score, r
    return best_row


def cell_text(v):
    if v is None:
        return ""
    return str(v).strip()


def split_list_cell(text):
    """측정값 cells are often comma/newline/slash separated lists."""
    if not text:
        return []
    parts = re.split(r"[,\n/、]+", text)
    return [p.strip() for p in parts if p.strip()]


def sql_hash(sql_text):
    return hashlib.sha1(norm(sql_text).encode("utf-8")).hexdigest()[:12]


def build_merge_map(ws):
    """Merged cells (e.g. one GOLD query spanning all rows of a visual
    group) read as value-in-top-left, None everywhere else in read_only
    mode. Map every covered cell -> its anchor cell so we can forward-fill."""
    merge_map = {}
    for rng in ws.merged_cells.ranges:
        anchor = (rng.min_row, rng.min_col)
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                merge_map[(row, col)] = anchor
    return merge_map


def main():
    # non-read-only: need ws.merged_cells.ranges (unreliable in read_only mode)
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True, read_only=False)

    workbook_inventory = {"workbook": SRC_XLSX.name, "sheets": []}
    visual_catalog = []
    measure_index = {}  # measure_name -> set of row_ids (as list) + role
    sql_registry = {}   # sql_hash -> {sql_id, sql_text, first_seen, referenced_by}
    unresolved = []

    sql_counter = 0

    for ws in wb.worksheets:
        sheet_name = ws.title
        merge_map = build_merge_map(ws)

        def cv(r, c):
            """Cell value, resolving merged ranges to their anchor value."""
            ar, ac = merge_map.get((r, c), (r, c))
            return ws.cell(row=ar, column=ac).value

        header_row = find_header_row(ws)
        headers = [cv(header_row, c) for c in range(1, ws.max_column + 1)]
        col_field = {}
        query_cols = []  # (col_index, header_label, priority) for query_candidate columns
        for c, h in enumerate(headers, start=1):
            f = classify_header(h)
            if f == "query_candidate":
                query_cols.append((c, cell_text(h), query_priority(h)))
            elif f and f not in col_field.values():
                col_field[c] = f
        query_cols.sort(key=lambda x: -x[2])
        gold_col = query_cols[0][0] if query_cols else None
        gold_col_label = query_cols[0][1] if query_cols else None
        alt_query_cols = query_cols[1:]  # (col_index, header_label, priority)

        if not col_field and gold_col is None:
            unresolved.append({
                "status": "unresolved",
                "reason": "헤더 행에서 알려진 컬럼 패턴을 하나도 인식하지 못함",
                "source": {"workbook": SRC_XLSX.name, "sheet": sheet_name, "row": header_row},
                "required_input": ["시트 레이아웃 확인 필요 (헤더 라벨 상이)"],
            })

        row_count = 0
        visual_count = 0
        metric_names_in_sheet = set()
        gold_query_count = 0

        for r in range(header_row + 1, ws.max_row + 1):
            raw = {}
            has_any = False
            for c in range(1, ws.max_column + 1):
                v = cv(r, c)
                if v is not None and str(v).strip() != "":
                    has_any = True
                if c == gold_col:
                    continue  # handled separately below
                alt = next((label for (ci, label, _p) in alt_query_cols if ci == c), None)
                if alt:
                    if cell_text(v):
                        raw[f"gold_query_variant__{norm(alt)}"] = cell_text(v)
                    continue
                field = col_field.get(c)
                key = field if field else f"col_{c}"
                raw[key] = cell_text(v)
            if not has_any:
                continue

            row_count += 1
            row_id = f"{sheet_name}_ROW_{r:03d}"

            visual_name = raw.get("visual_name", "")
            visual_type = raw.get("visual_type", "")
            value_label = raw.get("value_label", "")
            dax_expression = raw.get("dax_expression", "")
            used_measures = split_list_cell(raw.get("used_measures", ""))
            related_measures = split_list_cell(raw.get("related_measures", ""))
            gold_query = cell_text(cv(r, gold_col)) if gold_col else ""
            business_meaning = raw.get("business_meaning", "")
            slicers_filters = raw.get("slicers_filters", "")

            if visual_name or visual_type:
                visual_count += 1

            gold_query_id = None
            if gold_query and len(gold_query) > 15:  # skip stray short text mistakenly caught
                gold_query_count += 1
                h = sql_hash(gold_query)
                if h not in sql_registry:
                    sql_counter += 1
                    sql_registry[h] = {
                        "sql_id": f"gold_{sheet_name.split('.')[0].replace(' ', '').replace('-', '_')}_{sql_counter:03d}",
                        "source_hash": h,
                        "sql_text": gold_query,
                        "first_seen_row_id": row_id,
                        "referenced_by": [],
                    }
                sql_registry[h]["referenced_by"].append(row_id)
                gold_query_id = sql_registry[h]["sql_id"]

            for m in used_measures:
                metric_names_in_sheet.add(m)
                measure_index.setdefault(m, {"measure_name": m, "used_in": [], "related_in": []})
                measure_index[m]["used_in"].append(row_id)
            for m in related_measures:
                metric_names_in_sheet.add(m)
                measure_index.setdefault(m, {"measure_name": m, "used_in": [], "related_in": []})
                measure_index[m]["related_in"].append(row_id)

            entry = {
                "visual_id": row_id,
                "page_name": raw.get("page", ""),
                "visual_name": visual_name,
                "value_label": value_label,
                "visual_type": visual_type,
                "business_purpose": business_meaning,
                "used_measures": used_measures,
                "related_measures": related_measures,
                "dax_expression": dax_expression or None,
                "dimensions": [],  # filled in P0-2 (sql static analysis) from gold_query_id join/group_by
                "filters": [],
                "slicers_raw": slicers_filters,
                "date_conditions": None,
                "gold_query_id": gold_query_id,
                "gold_query_column_used": gold_col_label,
                "gold_query_variants_present": [label for (_ci, label, _p) in alt_query_cols
                                                 if raw.get(f"gold_query_variant__{norm(label)}")],
                "raw_columns": {k: v for k, v in raw.items()
                                if (k.startswith("col_") or k.startswith("gold_query_variant__")) and v},
                "source": {
                    "workbook": SRC_XLSX.name,
                    "sheet": sheet_name,
                    "row": r,
                    "visual_name": visual_name or None,
                    "measure_name": used_measures[0] if used_measures else None,
                    "sql_reference": gold_query_id,
                },
            }
            visual_catalog.append(entry)

            if not gold_query and (used_measures or related_measures):
                unresolved.append({
                    "status": "unresolved",
                    "reason": "측정값은 참조되었으나 해당 행에 GOLD 쿼리 원문이 없음",
                    "source": entry["source"],
                    "required_input": ["GOLD 쿼리 원문 또는 DAX 정의"],
                })

        workbook_inventory["sheets"].append({
            "sheet_name": sheet_name,
            "header_row": header_row,
            "row_count": row_count,
            "visual_count": visual_count,
            "metric_count": len(metric_names_in_sheet),
            "gold_query_count": gold_query_count,
            "gold_query_column_used": gold_col_label,
            "gold_query_alt_columns": [label for (_ci, label, _p) in alt_query_cols],
            "status": "parsed" if col_field else "unresolved",
        })

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    (OUT_DIR / "workbook_inventory.json").write_text(
        json.dumps(workbook_inventory, ensure_ascii=False, indent=2), encoding="utf-8")

    (OUT_DIR / "visual_catalog.json").write_text(
        json.dumps(visual_catalog, ensure_ascii=False, indent=2), encoding="utf-8")

    measure_inventory = sorted(measure_index.values(), key=lambda x: x["measure_name"])
    (OUT_DIR / "measure_inventory.json").write_text(
        json.dumps(measure_inventory, ensure_ascii=False, indent=2), encoding="utf-8")

    sql_inventory = sorted(sql_registry.values(), key=lambda x: x["sql_id"])
    (OUT_DIR / "sql_inventory.json").write_text(
        json.dumps(sql_inventory, ensure_ascii=False, indent=2), encoding="utf-8")

    (OUT_DIR / "unresolved_items.json").write_text(
        json.dumps(unresolved, ensure_ascii=False, indent=2), encoding="utf-8")

    # table_column_usage.json is populated by parse_gold_sql.py (P0 step 2, sqlglot AST pass)
    if not (OUT_DIR / "table_column_usage.json").exists():
        (OUT_DIR / "table_column_usage.json").write_text("[]", encoding="utf-8")

    summary = {
        "sheets": len(workbook_inventory["sheets"]),
        "total_rows": sum(s["row_count"] for s in workbook_inventory["sheets"]),
        "total_visuals": len(visual_catalog),
        "distinct_measures": len(measure_inventory),
        "distinct_gold_sql": len(sql_inventory),
        "unresolved_count": len(unresolved),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
